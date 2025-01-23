
from ctypes import alignment
from datetime import datetime
import json
import os
import random
from fastapi import HTTPException
import pandas as pd
import yaml
import logging
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.ERROR)
handler = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)
import json
import pandas as pd
import logging

class JSONToExcelConverter:
    # def convert_json_to_excel(self, final_result, excel_output_path, config_type):
    #     print("final result",final_result)
       
    #     try:
    #         # Convert dictionary to DataFrame
    #         timestamp = final_result[0]['results']['timestamp']
    #         print("final result",final_result)
            
    #         for model in final_result:  # Iterate over each model in Final_result
    #             model_id = model['model_id']  # Extract model_id
    #             final_result = model['results']  # Get the results for this model

    #             print("142686786")
    #             # Prepare data for the 'Raw Data' sheet
    #             raw_data = []
    #             for key, result in final_result.items():
    #                 print("hcghdhghsdg")

    #                 if key not in ["timestamp", "results"]:
    #                     if isinstance(result, list):
    #                         num_inputs = len({obj['query'] for obj in result})
    #                         for obj in result:
    #                             if 'test_id' in obj:
    #                                 existing_details = obj.get('test_id', '')
    #                                 new_details = f"(input={num_inputs}, config_id={model_id}, config_type={config_type})"
    #                                 obj['test_id'] = existing_details + (new_details if existing_details else new_details)
    #                             obj['Checkbox'] = random.choice(['Yes', 'No'])
    #                             raw_data.append(obj)

    #             df_raw_sheet = pd.DataFrame(raw_data)
    #             df_raw_sheet.insert(0, 'timestamp', timestamp)
    #             print("v4vd5d5v5")

    #             # Prepare data for 'Organized Data' sheet
    #             organized_data = []
    #             for key, result in final_result.items():
    #                 if key not in ["timestamp", "results"]:
    #                     test_ids = {obj['test_id'] for obj in result}
    #                     for test_id in test_ids:
    #                         queries = [q['query'] for q in result if q['test_id'] == test_id]
    #                         organized_data.append([f"test_id:{test_id} ", queries[0]])
    #                         for query in queries[1:]:
    #                             organized_data.append(["", query])

    #             columns_organized = ['test_id and details', 'Input']
    #             df_organized_sheet = pd.DataFrame(organized_data, columns=columns_organized)
    #             df_organized_sheet.insert(0, "timestamp", timestamp)
    #             print("-----",excel_output_path)
    #             # Check if the Excel file exists
    #             file_exists = os.path.isfile(excel_output_path)

    #             with pd.ExcelWriter(excel_output_path, engine='openpyxl', mode='a' if file_exists else 'w') as writer:
    #                 # Define a function to append with a gap
    #                 def append_with_gap(df, sheet_name, gap=2):
    #                     if file_exists and sheet_name in writer.sheets:
    #                         # Load the existing sheet
    #                         sheet = writer.sheets[sheet_name]
    #                         # Find the last row with data
    #                         last_row = sheet.max_row
    #                         # Insert gap rows
    #                         for _ in range(gap):
    #                             sheet.append([None] * df.shape[1])  # Add empty rows
    #                         # Append new data
    #                         for r in dataframe_to_rows(df, index=False, header=False):
    #                             sheet.append(r)
    #                     else:
    #                         # If the sheet does not exist, create it
    #                         df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

    #                 # If the file is new, write initial sheets
    #                 if not file_exists:
    #                     df_organized_sheet.to_excel(writer, sheet_name='Response Details', index=False, header=True)
    #                     df_raw_sheet.to_excel(writer, sheet_name='Input Details', index=False, header=True)
    #                 else:
    #                     # Append data with gaps for existing sheets
    #                     append_with_gap(df_organized_sheet, 'Response Details')
    #                     append_with_gap(df_raw_sheet, 'Input Details')
    #         print("excel output path........",excel_output_path)
    #         return {
    #                 "status_code": 200,
    #                 "path": excel_output_path,
    #                 "detail": "Appended results to the Excel file."
    #             }

    #     except ValueError as ve:
    #         return {
    #             "status_code": 400,
    #             "detail": str(ve),
    #         }
    #     except Exception as e:
    #         return {
    #             "status_code": 500,
    #             "detail": f"Failed to convert JSON to Excel: {str(e)}",
    #         }

    def convert_json_to_excel(self, final_result, excel_output_path, config_type):
        try:
            print("IN EXCEL...")
            # Ensure final_result is a dictionary
            if isinstance(final_result, dict):
                # Extract timestamp
                timestamp = final_result.get('timestamp')
            elif isinstance(final_result, list):
                # Handle the case where final_result is a list
                if len(final_result) > 0 and isinstance(final_result[0], dict):
                    timestamp = final_result[0].get('timestamp')
                else:
                    raise ValueError("Invalid structure in final_result: expected list of dictionaries with 'timestamp'.")
            else:
                raise ValueError("Invalid structure: final_result should be a dictionary or a list of dictionaries.")
            print("final result", final_result)
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(excel_output_path), exist_ok=True)

            # Generate timestamp for file naming
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            print("111")
            # Create a new file name with the timestamp
            file_name = f"result_{timestamp}.xlsx"
            print("file name", file_name)

            # Combine the file name with the directory path
            file_path = os.path.join(excel_output_path, file_name)
            print("Excel file path:", file_path)

            raw_data = []
            organized_data = []

            for result in final_result:  # Iterate over the list
                model_id = result.get('model_id')  # Extract 'model_id'
                results_payload = result.get('results', {})  # Extract 'results' as a dictionary

                # Check if there is a 'timestamp' in the results
                if 'timestamp' in results_payload:
                    timestamp = results_payload['timestamp']

                for key, payload in results_payload.items():
                    if key == 'timestamp':
                        continue  # Skip 'timestamp' as it is not part of the payloads

                    num_inputs = len(payload)  # Number of inputs in the payload

                    for obj in payload:  # Iterate over each payload object
                        # Prepare raw data
                        if 'test_id' in obj:
                            existing_details = obj.get('test_id', '')
                            new_details = f"(input={num_inputs}, config_id={model_id}, config_type={config_type})"
                            obj['test_id'] = existing_details + (new_details if existing_details else new_details)

                        obj['Checkbox'] = random.choice(['Yes', 'No'])  # Randomize checkbox value
                        raw_data.append(obj)

                        # Prepare organized data for Excel sheet
                        organized_data.append([f"test_id:{obj['test_id']} ", obj['query']])

            # Create DataFrames
            df_raw_sheet = pd.DataFrame(raw_data)
            df_raw_sheet.insert(0, 'timestamp', timestamp)

            columns_organized = ['test_id and details', 'Input']
            df_organized_sheet = pd.DataFrame(organized_data, columns=columns_organized)
            df_organized_sheet.insert(0, "timestamp", timestamp)

            # Write to Excel
            file_exists = os.path.isfile(file_path)
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a' if file_exists else 'w') as writer:
                if not file_exists:
                    df_organized_sheet.to_excel(writer, sheet_name='Response Details', index=False, header=True)
                    df_raw_sheet.to_excel(writer, sheet_name='Input Details', index=False, header=True)
                else:
                    # Append data with gaps
                    def append_with_gap(df, sheet_name, gap=2):
                        if sheet_name in writer.sheets:
                            sheet = writer.sheets[sheet_name]
                            for _ in range(gap):
                                sheet.append([None] * df.shape[1])
                            for r in dataframe_to_rows(df, index=False, header=False):
                                sheet.append(r)
                        else:
                            df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

                    append_with_gap(df_organized_sheet, 'Response Details')
                    append_with_gap(df_raw_sheet, 'Input Details')

            print("Excel output path:", file_path)
            return {
                "status_code": 200,
                "path": file_path,
                "detail": "Appended results to the Excel file."
            }


        except ValueError as ve:
            return {
                "status_code": 400,
                "detail": str(ve),
            }
        except Exception as e:
            return {
                "status_code": 500,
                "detail": f"Failed to convert JSON to Excel: {str(e)}",
            }

                
    def add_metrics_sheet(self, final_result, excel_output_path):
        try:

            # Check if the file exists and is a valid Excel file
            if not os.path.isfile(excel_output_path):
                raise FileNotFoundError(f"File not found: {excel_output_path}")
        
             # Load the existing Excel workbook
            book = load_workbook(excel_output_path)

            # third sheet data
            score_data = []
            for key, value in final_result.items():
                if key.startswith("Payload"):
                    for payload in value:
                        score_data.append({
                            "answer": payload.get("answer", ""),
                            "response": payload.get("response", ""),
                            "score": payload.get("score", "")
                        })
            for item in score_data:
                if item['score'] == '':
                    df_score_sheet = pd.DataFrame(score_data, columns=['answer', 'response'])
                else:
                    df_score_sheet = pd.DataFrame(score_data, columns=['answer', 'response', 'score'])


            #Metrics Data sheet
            
            df_bert_scores = pd.DataFrame()
            df_cm = pd.DataFrame()
            for key, value in final_result.items():
                if key  == "BERT_score":

                    scores = final_result["BERT_score"] 
                    column_name = "Bert_Score" 

                    # Construct metrics for the Bert/Rouge score in one row
                    df_bert_scores = pd.DataFrame({
                        column_name : ["Precision", "Recall", "F1-Score"],
                        "": [
                            scores.get("precision", 0),
                            scores.get("recall", 0),
                            scores.get("f1", 0), 
                        ]
                    }).T

                    

                # elif key  ==  "ROUGE_score":
                #     scores = final_result["ROUGE_score"] 
                #     column_name = "ROUGE_score" 
                #     print(scores, column_name)

                #     # Construct metrics for the Bert/Rouge score in one row
                #     df_rouge_scores = pd.DataFrame({
                #         column_name : ["ROUGE-1 F1", "ROUGE-2 F1", "ROUGE-L F1"],
                #         "": [
                #             scores.get("ROUGE-1 F1", 0),
                #             scores.get("ROUGE-2 F1", 0),
                #             scores.get("ROUGE-L F1", 0), 
                #         ]
                #     }).T

                #     print("------",df_rouge_scores)

                #     df_metrics_sheet = pd.concat(df_rouge_scores)
                #     print("------",df_metrics_sheet)

                elif(key == "confusion_matrix"):
                    confusion_matrix = final_result["confusion_matrix"]
                    df_cm = pd.DataFrame(confusion_matrix)
                    
                    list_metrics = list(df_cm)
                    metrics_to_remove = {'Precision', 'Recall', 'f1-score'}

                    list_metrics = [x for x in list_metrics if x not in metrics_to_remove]

                    # list_metrics.append("Accuracy")

                    # Add accuracy as a separate row
                    # accuracy = final_result.get('Accuracy', '')
                    # accuracy_row = pd.DataFrame([[""] * (len(df_metrics.columns) - 1) + [accuracy]], columns=df_metrics.columns)

                    df_cm.insert(0, "Actual/Predicted", list_metrics)
                    df_cm = df_cm.reset_index(drop=True)                   

                else:
                    df_metrics_sheet = pd.DataFrame()

            df_metrics_sheet = pd.concat([df_bert_scores, df_cm])
            # Optional: Round float values to 2 decimal places
            df_metrics_sheet = df_metrics_sheet.round(2)

            # Replace NaN with empty string for cleaner Excel output
            df_metrics_sheet = df_metrics_sheet.fillna("")

            # # Prepare Bert score metrics data
            # accuracy = final_result.get("Accuracy", "")

           

            # # Construct Accuracy data in a separate row
            # df_accuracy = pd.DataFrame({
            #     "Accuracy": [round(accuracy, 2)]
            # }).T

            # Check if the Excel file exists
            file_exists = os.path.isfile(excel_output_path)

            with pd.ExcelWriter(excel_output_path, engine='openpyxl', mode='a' if file_exists else 'w') as writer:
                # Define a function to append with a gap
                def append_with_gap(df, sheet_name, gap=2):
                    # Check if the sheet exists
                    if sheet_name in writer.sheets:
                        # Load the existing sheet
                        sheet = writer.sheets[sheet_name]
                        # Find the last row with data
                        last_row = sheet.max_row
                        # Insert gap rows
                        for _ in range(gap):
                            sheet.append([None] * df.shape[1])  # Add empty rows
                        # Append new data
                        for r in dataframe_to_rows(df, index=False, header=False):
                            sheet.append(r)
                    else:
                        # If the sheet does not exist, create it
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

                # If the file is new, write initial sheets
                if not file_exists:
                    df_score_sheet.to_excel(writer, sheet_name='Scores Sheet', index=False, header=True)
                    df_metrics_sheet.to_excel(writer, sheet_name='Metrics Sheet', index=False, header=True)
                else:
                    # Append data with gaps for existing sheets
                    append_with_gap(df_score_sheet, 'Scores Sheet')
                    append_with_gap(df_metrics_sheet, 'Metrics Sheet')

            book.save(excel_output_path)

            print("Excel file successfully created.")

            return {
                "status_code": 200,
                "detail": "Excel file successfully created."
            }

        except ValueError as ve:
            return {
                "status_code": 400,
                "detail": str(ve),
            }
        except Exception as e:
            return {
                "status_code": 500,
                "detail": f"Failed to convert JSON to Excel: {str(e)}",
            }
        
    def filter_checked_rows(excel_path):
        try:
            
            # Load the workbook to preserve existing sheets
            book = load_workbook(excel_path)

            # Read the 'Input Details' sheet into a DataFrame
            df = pd.read_excel(excel_path, sheet_name='Input Details')

            # Check if 'Checkbox' column exists
            if 'Checkbox' not in df.columns:
                return {
                    "status_code": 404,
                    "detail": "Checkbox column not found in the Excel file.",
                }

            # Initialize an empty dictionary for storing the payload data
            payload_data = {}
            payload_count = 1  # This will track the payload number

            # Loop through each unique test_id
            for test_id in df['test_id'].unique():
                # Filter the dataframe for the current test_id
                test_df = df[df['test_id'] == test_id]

                # Ensure case-insensitive matching for 'Checkbox' column
                test_df['Checkbox'] = test_df['Checkbox'].str.lower()

                # Filter rows where Checkbox is 'yes'
                filtered_test_df = test_df[test_df['Checkbox'] == 'yes']

                # Create a key for the current payload
                payload_key = f"Payload_{payload_count}"
                payload_count += 1

                # If there are rows after filtering, populate the payload
                if not filtered_test_df.empty:
                    # Initialize an empty dictionary to store records for this payload
                    payload_records = {}

                    for idx in range(len(filtered_test_df)):
                        record = filtered_test_df.iloc[idx]  # Get the record by index
                        index_key = f"Index {idx + 1}"  # Start indices from 1
                        payload_records[index_key] = {
                            "query": record["query"],
                            "response": record["response"],
                            "answer": record["answer"]
                        }

                    # Store the records in the payload data using the payload key
                    payload_data[payload_key] = payload_records
                else:
                    # If no records were found, add "none" for this payload
                    payload_data[payload_key] = "none"
            # Optionally save the filtered dataframe to a new sheet in the Excel file
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                test_df.to_excel(writer, sheet_name='Filtered Results', index=False)
            
            return {
                "status_code": 200,
                "data": payload_data
            }
        except Exception as e:
            return {
                "status_code": 500,
                "detail": f"Failed to get data: {str(e)}",
            }

